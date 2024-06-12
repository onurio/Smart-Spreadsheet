import openpyxl


def find_table_boundaries(sheet):
    rows = sheet.max_row
    cols = sheet.max_column
    visited = set()

    def is_bordered(cell):
        return (
            cell.border.top.style
            or cell.border.bottom.style
            or cell.border.left.style
            or cell.border.right.style
        )

    def get_next_bordered_cell(start_row, start_col):
        for r in range(start_row, rows + 1):
            for c in range(start_col, cols + 1):
                if (r, c) in visited:
                    continue
                cell = sheet.cell(r, c)
                if is_bordered(cell):
                    return r, c
        return None, None

    def get_table_boundary(start_row, start_col):
        min_row, max_row = start_row, start_row
        min_col, max_col = start_col, start_col

        queue = [(start_row, start_col)]
        while queue:
            r, c = queue.pop(0)
            if (r, c) in visited:
                continue

            visited.add((r, c))
            min_row = min(min_row, r)
            max_row = max(max_row, r)
            min_col = min(min_col, c)
            max_col = max(max_col, c)

            for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                nr, nc = r + dr, c + dc
                if 1 <= nr <= rows and 1 <= nc <= cols:
                    neighbor = sheet.cell(nr, nc)
                    if is_bordered(neighbor) and (nr, nc) not in visited:
                        queue.append((nr, nc))

        return min_row, max_row, min_col, max_col

    boundaries = []
    start_row, start_col = 1, 1
    while True:
        r, c = get_next_bordered_cell(start_row, start_col)
        if r is None:
            break

        boundary = get_table_boundary(r, c)
        boundaries.append(boundary)

        start_row, start_col = r + 1, 1  # Move to the next row and reset column to 1

    return boundaries


def extract_tables(sheet, table_boundaries):
    tables = []
    for boundary in table_boundaries:
        table = []
        for row in sheet.iter_rows(
            min_row=boundary[0],
            max_row=boundary[1],
            min_col=boundary[2],
            max_col=boundary[3],
        ):
            table.append([cell.value for cell in row])
        tables.append(table)
    return tables


def process_file(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    table_boundaries = find_table_boundaries(sheet)
    tables = extract_tables(sheet, table_boundaries)
    return tables
